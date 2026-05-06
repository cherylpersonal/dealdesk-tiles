"""
One-time setup script — creates DealDesk_Master.xlsx.

After running this script:
  1. Open DealDesk_Master.xlsx in Excel
  2. File → Save As → Excel Macro-Enabled Workbook (.xlsm)
     Save it as DealDesk_Master.xlsm (same folder)
  3. Press Alt+F11 to open the VBA editor
  4. Right-click on VBAProject (DealDesk_Master.xlsm) → Import File
     Select DealDeskMacro.bas
  5. Close the VBA editor, go back to the Instructions sheet
  6. Insert → Shapes → draw a button, right-click → Assign Macro → ProcessDAF
  7. Run the SetupProtection macro ONCE:
       Alt+F11 → Immediate Window (Ctrl+G) → type:  SetupProtection  → Enter
  8. Go to Tools → VBAProject Properties → Protection tab
     Check "Lock project for viewing" and set a password
  9. Save and close
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
COST_SRC   = r"C:\Users\99202928\OneDrive - RAK Ceramics PJSC\Cheryl\Others\Deal Desk\Cost Sheet_05052026.xlsx"
OUTPUT     = os.path.join(BASE_DIR, "DealDesk_Master.xlsx")

# ── Tracker headers (must match AppendTracker order in VBA) ───────────────────
TRACKER_HEADERS = [
    "DAF Reference No.", "LOB", "Channel", "State",
    "Project Name", "Developer Name", "Dealer Name", "Zonal Coordinator",
    "Submitted Date", "DD Received Date", "6-Hr Due Date",
    "Response Date", "TAT", "SLA Met", "Quote Version",
    "Approval Level", "Approver Name", "Approval Status", "Approval Date",
    "List Value", "Deal Value", "Avg Discount %", "Deal Margin %",
]

# ── Styles ────────────────────────────────────────────────────────────────────
def bold_fill(hex_color):
    fill = PatternFill("solid", fgColor=hex_color)
    font = Font(bold=True)
    return font, fill

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Load cost data ────────────────────────────────────────────────────────────
df_cost = pd.read_excel(COST_SRC, header=1)   # row 2 (index 1) is header

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Instructions (first tab visible to users) ───────────────────────
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_instr.sheet_view.showGridLines = False

ws_instr.column_dimensions["B"].width = 72
ws_instr.row_dimensions[2].height = 30

ws_instr["B2"] = "Deal Desk — DAF Processor"
ws_instr["B2"].font = Font(size=18, bold=True, color="1F4E79")

ws_instr["B4"] = "How to use"
ws_instr["B4"].font = Font(bold=True, size=12)

steps = [
    "1.  Click the 'Process DAF' button on this sheet.",
    "2.  Browse to and select the filled DAF Excel file (.xlsx).",
    "3.  The macro calculates margins and writes a 'Calculations' sheet into the DAF file.",
    "4.  The Deal Desk Tracker sheet in this workbook is updated automatically.",
    "5.  Both files are saved automatically.",
    "",
    "NOTE: This workbook contains a protected cost sheet. Do not share with non-Deal Desk staff.",
    "      The cost data is hidden and password-protected.",
]
for i, step in enumerate(steps, 5):
    ws_instr[f"B{i}"] = step
    if i == 12:
        ws_instr[f"B{i}"].font = Font(italic=True, color="721C24")

ws_instr["B14"] = "▶  Process DAF"
ws_instr["B14"].font = Font(bold=True, size=14, color="FFFFFF")
ws_instr["B14"].fill = PatternFill("solid", fgColor="1F4E79")
ws_instr["B14"].alignment = Alignment(horizontal="center", vertical="center")
ws_instr.row_dimensions[14].height = 28
# (User will assign ProcessDAF macro to this cell or a shape button)

# ── Sheet 2: CostData (will be very-hidden after SetupProtection runs) ────────
ws_cost = wb.create_sheet("CostData")

hdr_font, hdr_fill = bold_fill("E8F0FE")
hdr_font = Font(bold=True)

for ci, col in enumerate(df_cost.columns, 1):
    cell = ws_cost.cell(row=1, column=ci, value=col)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border()
    ws_cost.column_dimensions[cell.column_letter].width = max(len(str(col)) + 4, 12)

for ri, row_data in df_cost.iterrows():
    for ci, val in enumerate(row_data, 1):
        ws_cost.cell(row=ri + 2, column=ci, value=val)

ws_cost.freeze_panes = "A2"

# ── Sheet 3: Deal Desk Tracker ────────────────────────────────────────────────
ws_tracker = wb.create_sheet("Deal Desk Tracker")

t_font = Font(bold=True, color="FFFFFF")
t_fill = PatternFill("solid", fgColor="1F4E79")
t_border = thin_border()

for ci, h in enumerate(TRACKER_HEADERS, 1):
    cell = ws_tracker.cell(row=1, column=ci, value=h)
    cell.font = t_font
    cell.fill = t_fill
    cell.border = t_border
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws_tracker.column_dimensions[cell.column_letter].width = max(len(h) + 2, 14)

ws_tracker.row_dimensions[1].height = 28
ws_tracker.freeze_panes = "A2"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"\nCreated: {OUTPUT}")
print()
print("Next steps:")
print("  1. Open DealDesk_Master.xlsx in Excel")
print("  2. File > Save As > Excel Macro-Enabled Workbook (.xlsm)")
print("  3. Alt+F11 > right-click VBAProject > Import File > DealDeskMacro.bas")
print("  4. Immediate Window (Ctrl+G): type SetupProtection and press Enter")
print("  5. Tools → VBAProject Properties → Protection → lock with password")
print("  6. Save and distribute only to Deal Desk staff")
